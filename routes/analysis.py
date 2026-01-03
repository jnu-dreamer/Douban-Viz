# routes/analysis.py
"""数据分析相关路由模块

包含数据大屏、导出功能等。
"""
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, send_file

analysis_bp = Blueprint('analysis', __name__)


@analysis_bp.route("/analysis")
def analysis():
    """数据分析大屏页面"""
    from app import repo
    
    # 1. 概览数据
    genre_list = repo.get_genre_statistics()
    country_labels, country_counts = repo.get_country_statistics()
    
    # 2. 评分与年份数据
    score_labels, score_counts = repo.get_score_distribution()
    year_labels, year_counts = repo.get_year_distribution()
    
    return render_template(
        "analysis.html",
        genre_data=genre_list,
        country_labels=country_labels,
        country_counts=country_counts,
        score_labels=score_labels,
        score_counts=score_counts,
        year_labels=year_labels,
        year_counts=year_counts
    )


@analysis_bp.route("/score")
def score():
    """重定向到分析页面的趋势 Tab"""
    return redirect(url_for('analysis.analysis') + '#trends')


@analysis_bp.route("/cluster")
def cluster_page():
    """重定向到分析页面的聚类 Tab"""
    return redirect(url_for('analysis.analysis') + '#ai-hub')


@analysis_bp.route("/export")
def export_data():
    """导出电影原始数据为 Excel"""
    import openpyxl
    from app import repo
    
    movies = repo.get_all_movies()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "豆瓣电影数据"
    
    headers = ["ID", "链接", "封面", "片名", "评分", "评价人数", "简介", "年份", "国家/地区", "类型", "导演", "主演"]
    ws.append(headers)
    
    for movie in movies:
        ws.append(movie)
    
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(out, as_attachment=True, download_name="douban_movies.xlsx")


@analysis_bp.route("/export/stats")
def export_stats():
    """导出统计报表为 Excel"""
    import openpyxl
    from app import repo
    
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # 1. 评分分布
    ws1 = wb.create_sheet("评分分布")
    ws1.append(["评分", "数量"])
    s_labels, s_counts = repo.get_score_distribution()
    for l, c in zip(s_labels, s_counts):
        ws1.append([float(l), c])
    
    # 2. 年份分布
    ws2 = wb.create_sheet("年份分布")
    ws2.append(["年份", "数量"])
    y_labels, y_counts = repo.get_year_distribution()
    for l, c in zip(y_labels, y_counts):
        ws2.append([l, c])
    
    # 3. 国家/地区分布
    ws3 = wb.create_sheet("国家地区分布")
    ws3.append(["国家/地区", "数量"])
    c_labels, c_counts = repo.get_country_statistics()
    for l, c in zip(c_labels, c_counts):
        ws3.append([l, c])
    
    # 4. 类型分布
    ws4 = wb.create_sheet("类型分布")
    ws4.append(["类型", "数量"])
    genres = repo.get_genre_statistics()
    for g in genres:
        ws4.append([g['name'], g['value']])
    
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="douban_stats_report.xlsx")


@analysis_bp.route("/export/graph")
def export_graph():
    """导出关系图谱数据为 Excel"""
    import openpyxl
    from app import repo
    from analysis.graph import GraphService
    
    service = GraphService(repo)
    data = service.build_graph(limit_nodes=9999)
    
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Nodes Sheet
    ws_nodes = wb.create_sheet("Nodes (人物)")
    ws_nodes.append(["Id", "Label", "Category (0=Director, 1=Actor)", "SymbolSize (Frequency)"])
    for node in data.get("nodes", []):
        ws_nodes.append([node["id"], node["name"], node["category"], node["symbolSize"]])
    
    # Edges Sheet
    ws_edges = wb.create_sheet("Edges (关系)")
    ws_edges.append(["Source", "Target"])
    for link in data.get("links", []):
        ws_edges.append([link["source"], link["target"]])
    
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="douban_graph_data.xlsx")


@analysis_bp.route("/export/keywords")
def export_keywords():
    """导出词云关键词为 Excel"""
    import jieba.analyse
    import openpyxl
    from app import repo
    
    text = repo.get_all_intro_text()
    tags = jieba.analyse.extract_tags(text, topK=500, withWeight=True, allowPOS=('n', 'nz', 'v', 'vn'))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "简介关键词 (Top 500)"
    ws.append(["关键词", "权重 (TF-IDF)", "排名"])
    
    for idx, (word, weight) in enumerate(tags, 1):
        ws.append([word, weight, idx])
    
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="douban_keywords.xlsx")
